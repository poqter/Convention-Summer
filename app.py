import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import os, re
import numpy as np


# ── 전역 상수 ────────────────────────────────────────────────
TABLE_SEQ = 0

EXCL_PAYMETHOD = "일시납"
EXCL_GROUP_PATTERN = r"연금성|저축성"
EXCL_STATUS_PATTERN = r"철회|해약|실효"

RATE_LT10 = 50
RATE_LIFE_10P = 80
RATE_NONLIFE_10P = 150


# ── 유틸 ────────────────────────────────────────────────────
def unique_sheet_name(wb, base, limit=31):
    name = str(base)[:limit] if base else "Sheet"
    if name not in wb.sheetnames:
        return name
    i = 2
    while True:
        suffix = f"_{i}"
        trunc = limit - len(suffix)
        cand = f"{name[:trunc]}{suffix}"
        if cand not in wb.sheetnames:
            return cand
        i += 1


def safe_table_name(base: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_]", "_", base)
    if not re.match(r"^[A-Za-z_]", name):
        name = f"tbl_{name}"
    return name[:254]


def header_idx(ws, name, default=None):
    for i in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=i).value == name:
            return i
    return default


def format_money(x):
    try:
        return f"{float(x):,.0f} 원"
    except Exception:
        return ""


def autosize_columns_fast(ws, df: pd.DataFrame, padding=4, max_width=45):
    """
    ✅ 기존 autosize_columns_full(전체 셀 스캔) 대신:
    - 헤더 길이 + 각 컬럼에서 대표 샘플(상위 30개) 기반으로만 너비 계산
    → 체감 속도 크게 개선
    """
    if df is None or df.empty:
        # 그래도 헤더는 맞춰줌
        for j, col in enumerate(df.columns, 1):
            letter = ws.cell(row=1, column=j).column_letter
            ws.column_dimensions[letter].width = min(max(len(str(col)) + padding, 10), max_width)
        return

    sample = df.head(30).astype(str)
    for j, col in enumerate(df.columns, 1):
        header_len = len(str(col))
        sample_max = sample[col].map(len).max() if col in sample.columns else 0
        width = min(max(header_len, sample_max) + padding, max_width)
        letter = ws.cell(row=1, column=j).column_letter
        ws.column_dimensions[letter].width = width


# ── 데이터 준비 (캐시) ───────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_df_from_bytes(file_bytes: bytes) -> pd.DataFrame:
    columns_needed = [
        "수금자명", "계약일", "보험사", "상품명", "납입기간",
        "초회보험료", "쉐어율", "납입방법", "상품군2", "계약상태"
    ]
    return pd.read_excel(BytesIO(file_bytes), usecols=columns_needed)


def exclude_contracts(df: pd.DataFrame):
    """
    제외: 일시납 / 연금성·저축성 / 철회·해약·실효
    """
    needed = {"납입방법", "상품군2", "계약상태"}
    if not needed.issubset(df.columns):
        return df.copy(), pd.DataFrame()

    tmp = df.copy()
    tmp["납입방법"] = tmp["납입방법"].astype(str).str.strip()
    tmp["상품군2"] = tmp["상품군2"].astype(str).str.strip()
    tmp["계약상태"] = tmp["계약상태"].astype(str).str.strip()

    is_lumpsum = tmp["납입방법"].str.contains(EXCL_PAYMETHOD, na=False)
    is_savings = tmp["상품군2"].str.contains(EXCL_GROUP_PATTERN, regex=True, na=False)
    is_bad_status = tmp["계약상태"].str.contains(EXCL_STATUS_PATTERN, regex=True, na=False)

    is_excluded = is_lumpsum | is_savings | is_bad_status
    return tmp[~is_excluded].copy(), tmp[is_excluded].copy()


def build_excluded_with_reason(exdf: pd.DataFrame) -> pd.DataFrame:
    base_cols = ["수금자명", "계약일자", "보험사", "상품명", "납입기간", "보험료", "납입방법", "제외사유"]
    if exdf is None or exdf.empty:
        return pd.DataFrame(columns=base_cols)

    tmp = exdf.copy()

    def reason_row(row):
        r = []
        if EXCL_PAYMETHOD in str(row.get("납입방법", "")): r.append("일시납")
        if re.search(EXCL_GROUP_PATTERN, str(row.get("상품군2", ""))): r.append("연금/저축성")
        stt = str(row.get("계약상태", ""))
        if "철회" in stt: r.append("철회")
        if "해약" in stt: r.append("해약")
        if "실효" in stt: r.append("실효")
        return " / ".join(r) if r else "제외 조건 미상"

    tmp["제외사유"] = tmp.apply(reason_row, axis=1)

    out = tmp[["수금자명", "계약일", "보험사", "상품명", "납입기간", "초회보험료", "납입방법", "제외사유"]].copy()
    out.rename(columns={"계약일": "계약일자", "초회보험료": "보험료"}, inplace=True)

    out["계약일자"] = pd.to_datetime(out["계약일자"], errors="coerce").dt.strftime("%Y-%m-%d")
    out["납입기간"] = out["납입기간"].apply(lambda x: f"{int(float(x))}년" if pd.notnull(x) else "")
    out["보험료"] = out["보험료"].map(lambda x: f"{x:,.0f} 원" if pd.notnull(x) else "")
    return out[base_cols]


def classify_insurance_type(ins_series: pd.Series) -> pd.Series:
    s = ins_series.astype(str).str.strip()
    is_nonlife = s.str.contains(r"손해|손보|화재|해상", regex=True, na=False)
    return np.where(is_nonlife, "손해보험", "생명보험")


@st.cache_data(show_spinner=False)
def compute_manager_score_cached(df_valid: pd.DataFrame) -> pd.DataFrame:
    df = df_valid.copy()
    df.rename(columns={"계약일": "계약일자", "초회보험료": "보험료"}, inplace=True)

    df["납입기간_num"] = pd.to_numeric(df["납입기간"], errors="coerce").fillna(0).astype(int)
    df["보험구분"] = classify_insurance_type(df["보험사"])

    df["환산율"] = np.select(
        [
            df["납입기간_num"] < 10,
            (df["납입기간_num"] >= 10) & (df["보험구분"] == "생명보험"),
            (df["납입기간_num"] >= 10) & (df["보험구분"] == "손해보험"),
        ],
        [RATE_LT10, RATE_LIFE_10P, RATE_NONLIFE_10P],
        default=0
    ).astype(int)

    df["쉐어율"] = df["쉐어율"].apply(lambda x: float(str(x).replace("%", "")) if pd.notnull(x) else x)
    df["실적보험료"] = pd.to_numeric(df["보험료"], errors="coerce").fillna(0)
    df["환산금액"] = df["실적보험료"] * df["환산율"] / 100
    df["계약일자_raw"] = pd.to_datetime(df["계약일자"], errors="coerce")

    return df


def build_group_and_top3(df: pd.DataFrame):
    group = df.groupby("수금자명", dropna=False).agg(
        건수=("수금자명", "size"),
        실적보험료합계=("실적보험료", "sum"),
        환산금액합계=("환산금액", "sum"),
    ).reset_index()

    group["환산금액순위"] = group["환산금액합계"].rank(method="dense", ascending=False).astype(int)
    group["건수순위"] = group["건수"].rank(method="dense", ascending=False).astype(int)

    group = group[["환산금액순위", "건수순위", "수금자명", "건수", "실적보험료합계", "환산금액합계"]]
    group = group.sort_values(["환산금액순위", "건수순위", "수금자명"]).reset_index(drop=True)

    top_amt = group[group["환산금액순위"] <= 3].copy()
    top_amt = top_amt.sort_values(["환산금액순위", "건수순위", "수금자명"])
    top_amt = top_amt[["환산금액순위", "수금자명", "환산금액합계", "건수"]]

    top_cnt = group[group["건수순위"] <= 3].copy()
    top_cnt = top_cnt.sort_values(["건수순위", "환산금액순위", "수금자명"])
    top_cnt = top_cnt[["건수순위", "수금자명", "건수", "환산금액합계"]]

    return group, top_amt, top_cnt


def to_styled(df: pd.DataFrame) -> pd.DataFrame:
    _ = df.copy()
    _["계약일자"] = pd.to_datetime(_["계약일자"], errors="coerce").dt.strftime("%Y-%m-%d")
    _["납입기간"] = _["납입기간_num"].astype(int).astype(str) + "년"
    _["보험료"] = pd.to_numeric(_["보험료"], errors="coerce").fillna(0).map("{:,.0f} 원".format)
    _["쉐어율"] = _["쉐어율"].astype(str) + " %"
    _["실적보험료"] = _["실적보험료"].map("{:,.0f} 원".format)
    _["환산율"] = _["환산율"].astype(str) + " %"
    _["환산금액"] = _["환산금액"].map("{:,.0f} 원".format)

    return _[
        ["수금자명","계약일자","보험사","보험구분","상품명",
         "납입기간","보험료","쉐어율","실적보험료","환산율","환산금액"]
    ]


def sums(df: pd.DataFrame):
    return float(df["실적보험료"].sum()), float(df["환산금액"].sum())


# ── 엑셀 ────────────────────────────────────────────────────
def write_table(ws, df_for_sheet: pd.DataFrame, start_row: int = 1, name_suffix: str = "A"):
    global TABLE_SEQ

    r_idx = start_row - 1
    for r_idx, row in enumerate(dataframe_to_rows(df_for_sheet, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    end_col_letter = ws.cell(row=start_row, column=df_for_sheet.shape[1]).column_letter
    last_row = r_idx if df_for_sheet.shape[0] > 0 else start_row

    TABLE_SEQ += 1
    display_name = safe_table_name(f"tbl_{ws.title}_{name_suffix}_{TABLE_SEQ}")
    table = Table(displayName=display_name, ref=f"A{start_row}:{end_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    autosize_columns_fast(ws, df_for_sheet, padding=5)
    return last_row


def totals_block(ws, perf, score, start_row: int):
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    fill = PatternFill("solid", fgColor="F2F2F2")

    col_rate = header_idx(ws, "환산율", 1)
    col_perf = header_idx(ws, "실적보험료", 2)
    col_score = header_idx(ws, "환산금액", 3)

    row = start_row + 2
    ws.cell(row=row, column=col_rate, value="총 합계").alignment = Alignment(horizontal="center")

    c1 = ws.cell(row=row, column=col_perf, value=f"{perf:,.0f} 원")
    c2 = ws.cell(row=row, column=col_score, value=f"{score:,.0f} 원")
    for c in (c1, c2):
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for c in [col_rate, col_perf, col_score]:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = thin_border

    return row


def build_workbook(df: pd.DataFrame, group: pd.DataFrame, excluded_disp_all: pd.DataFrame,
                   top_amt: pd.DataFrame, top_cnt: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"

    r = 1
    ws.cell(row=r, column=1, value="환산금액합계 TOP3").font = Font(bold=True)
    r = write_table(ws, top_amt, start_row=r + 1, name_suffix="TOPAMT") + 2

    ws.cell(row=r, column=1, value="건수 TOP3").font = Font(bold=True)
    r = write_table(ws, top_cnt, start_row=r + 1, name_suffix="TOPCNT") + 2

    ws.cell(row=r, column=1, value="수금자별 요약(순위 포함)").font = Font(bold=True)
    summary_fmt = group.copy()
    summary_fmt["실적보험료합계"] = summary_fmt["실적보험료합계"].map(format_money)
    summary_fmt["환산금액합계"] = summary_fmt["환산금액합계"].map(format_money)
    r = write_table(ws, summary_fmt, start_row=r + 1, name_suffix="SUM") + 1

    if not excluded_disp_all.empty:
        ws.cell(row=r + 1, column=1, value="제외 계약 목록").font = Font(bold=True)
        _ = write_table(ws, excluded_disp_all, start_row=r + 2, name_suffix="EXC")

    collectors = sorted(df["수금자명"].astype(str).unique().tolist())
    for collector in collectors:
        sub = df[df["수금자명"].astype(str) == collector].copy()
        ws2 = wb.create_sheet(title=unique_sheet_name(wb, collector))

        styled_sub = to_styled(sub)
        last_row = write_table(ws2, styled_sub, start_row=1, name_suffix="NORM")

        # 금액 컬럼 최소 너비 (고정)
        for header in ["실적보험료", "환산금액"]:
            idx = header_idx(ws2, header)
            if idx:
                col_letter = ws2.cell(row=1, column=idx).column_letter
                cur = ws2.column_dimensions[col_letter].width
                ws2.column_dimensions[col_letter].width = 20 if (cur is None or cur < 20) else cur

        perf, score = sums(sub)
        next_row = totals_block(ws2, perf, score, start_row=last_row)

        ex_sub = excluded_disp_all[excluded_disp_all["수금자명"].astype(str) == collector]
        if not ex_sub.empty:
            ws2.cell(row=next_row + 2, column=1, value="제외 계약").font = Font(bold=True)
            write_table(ws2, ex_sub, start_row=next_row + 3, name_suffix="EXC")

    return wb


# ── 메인 ────────────────────────────────────────────────────
def run():
    st.set_page_config(page_title="매니저 업적 환산기", layout="wide")

    with st.sidebar:
        st.header("🧭 사용 방법")
        st.markdown(
            """
            **🖥️ 한화라이프랩 전산**  
            **- 📂 계약관리**  
            **- 📑 보유계약 장기**  
            **- ⏱️ 기간 설정**  
            **- 💾 엑셀 다운로드 후 파일 첨부**
            """
        )
        st.divider()
        st.markdown(
            f"""
            **📌 환산 기준**  
            - 10년납 미만: **{RATE_LT10}%**  
            - 10년납 이상(생명): **{RATE_LIFE_10P}%**  
            - 10년납 이상(손해): **{RATE_NONLIFE_10P}%**
            """
        )
        st.markdown("**🚫 제외 기준**  \n- 일시납 / 연금성·저축성 / 철회·해약·실효")

    st.title("🏆 매니저 업적 환산기")
    st.caption("여러 명 선택 가능 · 선택된 수금자만 합산 결과/요약/엑셀로 출력합니다.")

    uploaded_file = st.file_uploader("📂 계약 목록 Excel 파일 업로드 (.xlsx)", type=["xlsx"])
    if not uploaded_file:
        st.info("📤 계약 목록 Excel 파일(.xlsx)을 업로드해주세요.")
        return

    file_bytes = uploaded_file.getvalue()
    base_filename = os.path.splitext(uploaded_file.name)[0]
    download_filename = f"{base_filename}_매니저업적_환산결과.xlsx"

    raw = load_df_from_bytes(file_bytes)

    df_valid, excluded_df = exclude_contracts(raw)
    excluded_disp_all = build_excluded_with_reason(excluded_df)

    # 필수 컬럼 체크
    df_valid.rename(columns={"계약일": "계약일자", "초회보험료": "보험료"}, inplace=True)
    required_columns = {"수금자명", "계약일자", "보험사", "상품명", "납입기간", "보험료", "쉐어율"}
    if not required_columns.issubset(df_valid.columns):
        st.error("❌ 업로드된 파일에 다음 항목이 모두 포함되어 있어야 합니다:\n" + ", ".join(sorted(required_columns)))
        st.stop()
    if df_valid["쉐어율"].isnull().any():
        st.error("❌ '쉐어율'에 빈 값이 포함되어 있습니다. 모든 행에 값을 입력해주세요.")
        st.stop()

    df_all = compute_manager_score_cached(df_valid)

    # 날짜 경고
    invalid_dates = df_all[df_all["계약일자_raw"].isna()]
    if not invalid_dates.empty:
        st.warning(f"⚠️ {len(invalid_dates)}건의 계약일자가 날짜로 인식되지 않았습니다. 엑셀에서 '2025-07-23'처럼 입력해주세요.")

    # 제외 건 표시
    if not excluded_df.empty:
        st.warning(f"⚠️ 제외된 계약 {len(excluded_df)}건 (일시납 / 연금성·저축성 / 철회·해약·실효)")
        with st.expander("🚫 제외된 계약 목록 보기"):
            excluded_display = excluded_df[["수금자명","계약일","보험사","상품명","납입기간","초회보험료","납입방법","계약상태","상품군2"]].copy()
            excluded_display.rename(columns={"초회보험료":"보험료"}, inplace=True)
            st.dataframe(excluded_display, use_container_width=True)

    # 멀티선택
    all_collectors = sorted(df_all["수금자명"].astype(str).unique().tolist())
    col1, col2 = st.columns([1, 2])
    with col1:
        use_all = st.checkbox("전체 선택", value=True)
    with col2:
        default_sel = all_collectors if use_all else (all_collectors[:1] if all_collectors else [])
        selected = st.multiselect("👤 수금자명 여러 명 선택(선택된 사람만 합산)", all_collectors, default=default_sel)

    if not selected:
        st.warning("선택된 수금자가 없습니다. 1명 이상 선택해주세요.")
        return

    show_df = df_all[df_all["수금자명"].astype(str).isin(selected)].copy()

    st.subheader("📄 선택된 수금자 합산 기준 환산 결과")
    st.dataframe(to_styled(show_df), use_container_width=True)

    perf_sum, score_sum = sums(show_df)
    st.subheader("📈 총합")
    st.markdown(
        f"""
        <div style='border:2px solid #1f77b4;border-radius:10px;padding:16px;background:#f7faff;'>
            <h4 style='color:#1f77b4;margin:0;'>📈 총합 요약</h4>
            <p style='margin:6px 0;'><strong>▶ 실적보험료 합계:</strong> {perf_sum:,.0f} 원</p>
            <p style='margin:6px 0;'><strong>▶ 환산금액 합계:</strong> {score_sum:,.0f} 원</p>
            <p style='margin:6px 0;'><strong>▶ 선택 수금자:</strong> {len(selected)}명</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.subheader("🧮 수금자별 요약(순위 포함)")
    group, top_amt, top_cnt = build_group_and_top3(show_df)

    st.markdown("#### 🏅 환산금액합계 TOP3")
    top_amt_disp = top_amt.copy()
    top_amt_disp["환산금액합계"] = top_amt_disp["환산금액합계"].map(format_money)
    st.dataframe(top_amt_disp, use_container_width=True)

    st.markdown("#### 🏅 건수 TOP3")
    top_cnt_disp = top_cnt.copy()
    top_cnt_disp["환산금액합계"] = top_cnt_disp["환산금액합계"].map(format_money)
    st.dataframe(top_cnt_disp, use_container_width=True)

    disp_group = group.copy()
    disp_group["실적보험료합계"] = disp_group["실적보험료합계"].map(format_money)
    disp_group["환산금액합계"] = disp_group["환산금액합계"].map(format_money)
    st.dataframe(disp_group, use_container_width=True)

    wb = build_workbook(show_df, group, excluded_disp_all, top_amt, top_cnt)
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    st.download_button(
        label="📥 환산 결과 엑셀 다운로드 (TOP3 + 요약 + 수금자별 시트 + 제외사유)",
        data=out,
        file_name=download_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    run()
